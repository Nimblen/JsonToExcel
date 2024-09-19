import json
import pytest
import openpyxl
from openpyxl.styles import Alignment, Font
import os

from scripts.main import create_excel_from_data, create_json_from_excel, load_json_data

@pytest.fixture()
def test_data():
    TEST_DATA = {
        "employees": [
            {
                "ID": 637,
                "Name": "Rajesh",
                "Department": "Engineering",
                "Role": "Senior Engineer",
                "Salary": 95000,
                "Start Date": "2021-03-15",
                "Position": "Senior",
                "End Date": None,
                "Reason": None,
            },
        ],
        "offices": [
            {
                "ID": 1,
                "Name": "Office 1",
                "Location": "New York",
                "Employee ID": 637,
            },
        ],
    }
    TEST_EXCEL_FILE = "test.xlsx"
    TEST_JSON_FILE = "test.json"

    return TEST_DATA, TEST_EXCEL_FILE, TEST_JSON_FILE



def test_create_excel_from_data(test_data):
    TEST_DATA, TEST_EXCEL_FILE, _ = test_data
    create_excel_from_data(TEST_DATA, path_to_output_file_name=TEST_EXCEL_FILE)
    wb = openpyxl.load_workbook(TEST_EXCEL_FILE)
    assert wb.sheetnames[0] == "employees"
    assert wb.sheetnames[1] == "offices"
    
    ws = wb["employees"]
    assert ws["A1"].value == "ID"
    assert ws["A2"].value == 637

    ws = wb["offices"]
    assert ws["A1"].value == "ID"
    assert ws["A2"].value == 1

def test_create_json_from_excel(test_data):
    _, TEST_EXCEL_FILE, TEST_JSON_FILE = test_data
    create_json_from_excel(TEST_EXCEL_FILE, TEST_JSON_FILE)
    with open(TEST_JSON_FILE, "r") as f:
        data = json.load(f)
        assert data["employees"][0]["ID"] == 637
        assert data["offices"][0]["ID"] == 1

def test_load_json_data(test_data):
    _, _, TEST_JSON_FILE = test_data
    data = load_json_data(path_to_input_file=TEST_JSON_FILE)
    assert data["employees"][0]["ID"] == 637
    assert data["offices"][0]["ID"] == 1

def test_set_style(test_data):
    _, TEST_EXCEL_FILE, _ = test_data
    wb = openpyxl.load_workbook(TEST_EXCEL_FILE)
    ws_employees = wb["employees"]
    ws_offices = wb["offices"]

    ws_employees["A1"].alignment = Alignment(horizontal="center")
    assert ws_employees["A1"].alignment.horizontal == "center"

    ws_employees["A1"].alignment = Alignment(vertical="center")
    assert ws_employees["A1"].alignment.vertical == "center"

    ws_offices["A1"].font = Font(bold=True)
    assert ws_offices["A1"].font.bold is True

@pytest.fixture(autouse=True)
def setup_teardown(test_data):
    TEST_DATA, TEST_EXCEL_FILE, TEST_JSON_FILE = test_data

    if not os.path.exists(TEST_EXCEL_FILE):
        create_excel_from_data(TEST_DATA, path_to_output_file_name=TEST_EXCEL_FILE)
    if not os.path.exists(TEST_JSON_FILE):
        create_json_from_excel(TEST_EXCEL_FILE, TEST_JSON_FILE)

    yield

    if os.path.exists(TEST_EXCEL_FILE):
        os.remove(TEST_EXCEL_FILE)
    if os.path.exists(TEST_JSON_FILE):
        os.remove(TEST_JSON_FILE)

