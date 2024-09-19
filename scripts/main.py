import json
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from commands import parser

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
BORDER_STYLE = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
COLUMN_WIDTH_PADDING = 2
ARGS = parser.parse_args()

DEFAULT_PATH_TO_OUTPUT_FOLDER = "data/"


def style_cell(
    cell,
    alignment: Alignment = ALIGN_LEFT,
    font: Font = None,
    border: Border = BORDER_STYLE,
):
    cell.alignment = alignment
    cell.border = border
    if font:
        cell.font = font


def create_headers(ws, headers: list, column_mapping: dict):
    """Create headers for excel file(A1, B1, C1, ...)"""
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        # highlight headers
        style_cell(cell, alignment=ALIGN_CENTER, font=Font(bold=True))
        column_mapping[header] = col_letter


def write_row(ws, row, column_mapping, row_num):
    for col_name, col_letter in column_mapping.items():
        cell = ws[f"{col_letter}{row_num}"]
        cell.value = row.get(col_name, "")
        style_cell(cell)


def set_column_width(ws, column_mapping: dict, rows_data: list):
    for col_letter in column_mapping.values():
        max_length = max(
            len(str(ws[f"{col_letter}{row_num}"].value))
            for row_num in range(1, len(rows_data) + 1)
        )
        ws.column_dimensions[col_letter].width = max_length + COLUMN_WIDTH_PADDING


def create_excel_from_data(
    data: dict,
    path_to_output_file_name: str = f"{DEFAULT_PATH_TO_OUTPUT_FOLDER}output.xlsx",
):
    wb = openpyxl.Workbook()
    # del the default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, rows_data in data.items():
        ws = wb.create_sheet(title=sheet_name)
        column_mapping = {}
        all_columns = []

        for row in rows_data:
            for key in row.keys():
                if key not in all_columns:
                    all_columns.append(key)

        create_headers(ws, all_columns, column_mapping)

        for row_num, row in enumerate(rows_data, 2):
            write_row(ws, row, column_mapping, row_num)

        set_column_width(ws, column_mapping, rows_data)
    try:
        wb.save(path_to_output_file_name)
    except Exception as e:
        print(f"Error saving Excel file: {e}")


def create_json_from_excel(
    path_to_excel_file,
    path_to_output_file_name=f"{DEFAULT_PATH_TO_OUTPUT_FOLDER}output.json",
):
    wb = openpyxl.load_workbook(path_to_excel_file)
    result_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        table_data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            table_data.append(row_data)

        result_data[sheet_name] = table_data
    try:
        with open(
            path_to_output_file_name,
            "w",
            encoding="utf-8",
        ) as f:
            json.dump(result_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Error saving JSON data: {e}")


def load_json_data(path_to_input_file: str):
    try:
        with open(path_to_input_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading JSON data: {e}")
        return {}


if __name__ == "__main__":
    if ARGS.input:
        input_file = ARGS.input
        file_type = input_file.split(".")[-1]
        output_file = ARGS.output

        if file_type == "xlsx":
            create_json_from_excel(
                input_file,
                path_to_output_file_name=(
                    output_file
                    if output_file
                    else f"{DEFAULT_PATH_TO_OUTPUT_FOLDER}output.json"
                ),
            )
        elif file_type == "json":
            data = load_json_data(input_file)
            create_excel_from_data(
                data,
                path_to_output_file_name=(
                    output_file
                    if output_file
                    else f"{DEFAULT_PATH_TO_OUTPUT_FOLDER}output.xlsx"
                ),
            )
        else:
            print(f"Unsupported file type: {file_type}")
